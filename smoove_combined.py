"""
smoove_combined.py
------------------
מושך 4 רשימות הדרכה מ-Smoove, משווה לפי טלפון נייד,
ויוצר קובץ אקסל עם כל מי שחסר לפחות בהדרכה אחת.
"""

import os
import smtplib
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime

# ── הגדרות ───────────────────────────────────────────────────────────────────
SMOOVE_API_KEY    = os.environ["SEMOV_API_KEY"]
SMOOVE_BASE_URL   = "https://rest.smoove.io/v1"

# רשימות קובעות
FIXED_LIST_GENERAL = os.environ["FIXED_LIST_GENERAL"]   # בטיחות/שריפה/הטרדה: 1142366
FIXED_LIST_ACCESS  = os.environ["FIXED_LIST_ACCESS"]    # נגישות: 1100797

# רשימות משתנות (מי שכבר ביצע)
VAR_LIST_SAFETY    = os.environ["VAR_LIST_SAFETY"]      # בטיחות כללית: 1098539
VAR_LIST_FIRE      = os.environ["VAR_LIST_FIRE"]        # שריפות: 1098540
VAR_LIST_HARASS    = os.environ["VAR_LIST_HARASS"]      # הטרדה מינית: 1098541
VAR_LIST_ACCESS    = os.environ["VAR_LIST_ACCESS"]      # נגישות: 1098542

# רשימות תזכורת SMS
SMS_LIST_SAFETY   = os.environ["SMS_LIST_SAFETY"]    # תזכורת בטיחות: 1145080
SMS_LIST_FIRE     = os.environ["SMS_LIST_FIRE"]      # תזכורת שריפות: 1145081
SMS_LIST_HARASS   = os.environ["SMS_LIST_HARASS"]    # תזכורת הטרדה: 1145082
SMS_LIST_ACCESS   = os.environ["SMS_LIST_ACCESS"]    # תזכורת נגישות: 1145083

SMS_LIST_MAP = {
    "בטיחות כללית":          "SMS_LIST_SAFETY",
    "מניעת שריפות וכיבוי אש": "SMS_LIST_FIRE",
    "מניעת הטרדה מינית":      "SMS_LIST_HARASS",
    "נגישות":                 "SMS_LIST_ACCESS",
}

EMAIL_SENDER      = os.environ["EMAIL_SENDER"]
EMAIL_PASSWORD    = os.environ["EMAIL_PASSWORD"]
EMAIL_RECIPIENT   = os.environ["EMAIL_RECIPIENT"]
SMTP_HOST         = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT         = int(os.environ.get("SMTP_PORT", "587"))

TRAININGS = [
    {"name": "בטיחות כללית",          "fixed": FIXED_LIST_GENERAL, "variable": VAR_LIST_SAFETY},
    {"name": "מניעת שריפות וכיבוי אש", "fixed": FIXED_LIST_GENERAL, "variable": VAR_LIST_FIRE},
    {"name": "מניעת הטרדה מינית",      "fixed": FIXED_LIST_GENERAL, "variable": VAR_LIST_HARASS},
    {"name": "נגישות",                 "fixed": FIXED_LIST_ACCESS,  "variable": VAR_LIST_ACCESS},
]

# ── Smoove API ────────────────────────────────────────────────────────────────

def get_contacts(list_id):
    headers = {"Authorization": f"Bearer {SMOOVE_API_KEY}"}
    contacts, page = [], 1
    while True:
        resp = requests.get(
            f"{SMOOVE_BASE_URL}/Lists/{list_id}/Contacts",
            headers=headers,
            params={"page": page, "itemsPerPage": 100},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        batch = data if isinstance(data, list) else (data.get("contacts") or data.get("data") or [])
        if not batch:
            break
        contacts.extend(batch)
        if isinstance(data, list):
            if len(batch) < 100:
                break
            page += 1
        else:
            total = data.get("total_pages") or data.get("meta", {}).get("total_pages", 1)
            if page >= total:
                break
            page += 1
    print(f"  → {len(contacts)} קונטקטים ברשימה {list_id}")
    return contacts

def normalize_phone(raw):
    digits = "".join(c for c in str(raw) if c.isdigit())
    if digits.startswith("972") and len(digits) >= 11:
        digits = "0" + digits[3:]
    return digits

def make_key(contact):
    raw = contact.get("cellPhone") or contact.get("mobile") or contact.get("phone") or ""
    return normalize_phone(raw.strip())

def get_contact_info(contact):
    return {
        "first":   contact.get("firstName") or contact.get("first_name") or "",
        "last":    contact.get("lastName")  or contact.get("last_name")  or "",
        "company": contact.get("company")   or "",
        "phone":   contact.get("cellPhone") or contact.get("mobile") or contact.get("phone") or "",
    }

# ── הוספה לרשימות SMS ────────────────────────────────────────────────────────

def bulk_add_to_sms_list(phones, list_id):
    """מוסיף רשימת קונטקטים לרשימת תזכורת SMS ב-Smoove - Bulk Import."""
    headers = {
        "Authorization": f"Bearer {SMOOVE_API_KEY}",
        "Content-Type": "application/json",
    }
    contacts = [{"cellPhone": phone} for phone in phones]
    payload = {
        "contacts": contacts,
        "lists_ToSubscribe": [int(list_id)]
    }
    resp = requests.post(
        f"{SMOOVE_BASE_URL}/Contacts_BulkImport",
        headers=headers,
        json=payload,
        timeout=60,
    )
    if resp.status_code not in (200, 201, 204):
        print(f"  ⚠️ Bulk Import נכשל ({resp.status_code}): {resp.text[:200]}")
        return 0
    result = resp.json()
    inserted = result.get("importReport", {}).get("inserted", 0)
    updated  = result.get("importReport", {}).get("updated", 0)
    return inserted + updated

def send_sms_reminders(missing_people, all_contacts_by_key):
    """מוסיף חסרים לרשימות SMS ב-Smoove בשיטת Bulk Import."""
    print("📱 שולח תזכורות SMS...")
    sms_list_vars = {
        "בטיחות כללית":          os.environ["SMS_LIST_SAFETY"],
        "מניעת שריפות וכיבוי אש": os.environ["SMS_LIST_FIRE"],
        "מניעת הטרדה מינית":      os.environ["SMS_LIST_HARASS"],
        "נגישות":                 os.environ["SMS_LIST_ACCESS"],
    }

    # בנה רשימות טלפונים לכל הדרכה
    phones_per_list = {list_id: [] for list_id in sms_list_vars.values()}

    for phone_key, person in missing_people.items():
        phone = person.get("phone") or ""
        if not phone:
            continue
        for training in person["missing"]:
            list_id = sms_list_vars.get(training)
            if list_id:
                phones_per_list[list_id].append(phone)

    # שלח Bulk Import לכל רשימה
    total_sent = 0
    for list_id, phones in phones_per_list.items():
        if not phones:
            continue
        print(f"  → מוסיף {len(phones)} קונטקטים לרשימה {list_id}")
        # חלק ל-500 בכל פעם (מגבלת Smoove)
        for i in range(0, len(phones), 500):
            batch = phones[i:i+500]
            added = bulk_add_to_sms_list(batch, list_id)
            total_sent += added

    print(f"  -> סהכ {total_sent} הוספות לרשימות SMS")

# ── לוגיקה ───────────────────────────────────────────────────────────────────

def build_missing_report():
    """
    מחזיר tuple: (missing_dict, contacts_by_key)
    missing_dict: { phone_key -> {info, missing: []} }
    contacts_by_key: { phone_key -> raw_contact }
    """
    all_people = {}   # phone_key -> info
    contacts_by_key = {}  # phone_key -> raw contact

    # שלב 1: אסוף את כל האנשים מכל הרשימות הקובעות
    print("⏳ שולף רשימות קובעות...")
    fixed_general = get_contacts(FIXED_LIST_GENERAL)
    fixed_access  = get_contacts(FIXED_LIST_ACCESS)

    for c in fixed_general + fixed_access:
        key = make_key(c)
        if key and key not in all_people:
            all_people[key] = {**get_contact_info(c), "missing": []}
            contacts_by_key[key] = c

    print(f"  → סה\"כ {len(all_people)} אנשים ייחודיים")

    # שלב 2: לכל הדרכה – מצא מי חסר
    print("🔍 בודק חסרים לכל הדרכה...")
    for training in TRAININGS:
        variable = get_contacts(training["variable"])
        completed_keys = {make_key(c) for c in variable}

        # מי מהרשימה הקובעת הרלוונטית לא השלים?
        fixed = get_contacts(training["fixed"]) if training["fixed"] == FIXED_LIST_ACCESS else fixed_general
        if training["fixed"] == FIXED_LIST_ACCESS:
            fixed = fixed_access
        else:
            fixed = fixed_general

        for c in fixed:
            key = make_key(c)
            if key and key not in completed_keys:
                if key not in all_people:
                    all_people[key] = {**get_contact_info(c), "missing": []}
                all_people[key]["missing"].append(training["name"])

    # החזר רק מי שחסר לפחות הדרכה אחת
    missing = {k: v for k, v in all_people.items() if v["missing"]}
    return missing, contacts_by_key

# ── יצירת אקסל ───────────────────────────────────────────────────────────────

def create_excel(missing_people):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "חוסרים"
    ws.sheet_view.rightToLeft = True

    # סגנונות
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", start_color="1a73e8")
    center_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align    = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    thin_border    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    alt_fill       = PatternFill("solid", start_color="EEF4FF")

    # כותרות
    headers = ["שם פרטי", "שם משפחה", "חברה", "טלפון נייד", "הדרכות חסרות"]
    widths  = [15, 15, 20, 15, 50]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    # נתונים
    for row_idx, (_, person) in enumerate(sorted(missing_people.items(),
                                                  key=lambda x: x[1]["last"]), 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        values = [
            person["first"],
            person["last"],
            person["company"],
            person["phone"],
            ", ".join(person["missing"]),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = right_align
            cell.border    = thin_border
            if fill.fill_type:
                cell.fill = fill

    # הקפא שורת כותרת
    ws.freeze_panes = "A2"

    # סיכום בתחתית
    total_row = len(missing_people) + 3
    ws.cell(row=total_row, column=1, value=f"סה\"כ חסרים: {len(missing_people)} אנשים").font = Font(name="Arial", bold=True, size=10)

    path = "/tmp/דוח_חוסרים_כולל.xlsx"
    wb.save(path)
    print(f"  → קובץ אקסל נוצר: {path}")
    return path

# ── שליחת מייל ───────────────────────────────────────────────────────────────

def send_email(excel_path, count):
    today = datetime.now().strftime("%d/%m/%Y")
    subject = f"דוח חוסרים כולל – {today} ({count} אנשים)"

    msg = MIMEMultipart()
    recipients = [r.strip() for r in EMAIL_RECIPIENT.split(",")]
    msg["Subject"] = subject
    msg["From"]    = EMAIL_SENDER
    msg["To"]      = ", ".join(recipients)

    body = f"""שלום,

מצורף דוח חוסרים כולל לתאריך {today}.
סה"כ {count} אנשים שלא השלימו לפחות הדרכה אחת.

הדוח כולל: שם, חברה, טלפון, ואיזה הדרכות חסרות לכל אחד.

בברכה,
מערכת האוטומציה"""

    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    filename = f"report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_SENDER, recipients, msg.as_string())

    print(f"  → מייל נשלח עם קובץ אקסל אל {EMAIL_RECIPIENT}")

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("⏳ בונה דוח חוסרים כולל...")
    missing, contacts_by_key = build_missing_report()
    print(f"  → {len(missing)} אנשים עם לפחות הדרכה חסרת אחת")

    if not missing:
        print("✅ אין חוסרים – לא נשלח מייל ולא נשלחות הודעות SMS.")
        return

    print("📊 יוצר קובץ אקסל...")
    excel_path = create_excel(missing)

    print("📧 שולח מייל...")
    send_email(excel_path, len(missing))

    print("📱 שולח תזכורות SMS...")
    send_sms_reminders(missing, contacts_by_key)

    print("✅ סיום.")

if __name__ == "__main__":
    main()

